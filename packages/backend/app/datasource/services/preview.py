from __future__ import annotations

import csv
import io
import json
import math
import uuid
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

import pandas as pd
import pyarrow.parquet as pq
from openpyxl import load_workbook
from sqlalchemy import MetaData, Table, func, inspect, select

from app.core.config import settings
from app.datasource.services.specs import normalize_datasource_type, validate_database_datasource_type, validate_file_type
from app.infra.db import create_engine, normalize_connection_string
from app.infra.services.minio import download_bytes

DEFAULT_PREVIEW_PAGE_SIZE = 25
MAX_PREVIEW_PAGE_SIZE = 100


def build_datasource_preview(
    *,
    datasource,
    table_name: str | None = None,
    page: int = 1,
    page_size: int = DEFAULT_PREVIEW_PAGE_SIZE,
) -> dict[str, Any]:
    page = max(1, int(page))
    page_size = max(1, min(int(page_size), MAX_PREVIEW_PAGE_SIZE))
    category = getattr(datasource, "category", "database")

    if category == "database":
        return _build_database_preview(
            datasource=datasource,
            table_name=table_name,
            page=page,
            page_size=page_size,
        )
    if category == "file":
        return _build_file_preview(
            datasource=datasource,
            table_name=table_name,
            page=page,
            page_size=page_size,
        )
    raise ValueError(f"Unsupported datasource category: {category}")


def _build_database_preview(*, datasource, table_name: str | None, page: int, page_size: int) -> dict[str, Any]:
    datasource_type = normalize_datasource_type(getattr(datasource, "type", None))
    validate_database_datasource_type(datasource_type)

    connection_string = normalize_connection_string((getattr(datasource, "connection_string", None) or "").strip())
    if not connection_string:
        raise ValueError("Datasource has no connection_string")

    engine = create_engine(connection_string)
    try:
        inspector = inspect(engine)
        table_names = sorted(inspector.get_table_names())
        if not table_names:
            return _preview_payload(
                datasource=datasource,
                tables=[],
                selected_table=None,
                columns=[],
                rows=[],
                page=page,
                page_size=page_size,
                total_rows=0,
            )

        selected_table = table_name.strip() if table_name and table_name.strip() in table_names else table_names[0]
        reflected_table = Table(selected_table, MetaData(), autoload_with=engine)
        columns = [{"name": column.name, "type": str(column.type)} for column in reflected_table.columns]
        offset = (page - 1) * page_size

        with engine.connect() as connection:
            total_rows = int(connection.execute(select(func.count()).select_from(reflected_table)).scalar_one() or 0)
            rows_result = connection.execute(select(reflected_table).limit(page_size).offset(offset))
            rows = [_serialize_preview_row(dict(row)) for row in rows_result.mappings()]

        return _preview_payload(
            datasource=datasource,
            tables=[{"name": name} for name in table_names],
            selected_table=selected_table,
            columns=columns,
            rows=rows,
            page=page,
            page_size=page_size,
            total_rows=total_rows,
        )
    except Exception as exc:  # pragma: no cover - driver-specific failures vary
        raise ValueError(f"Failed to preview datasource: {exc}") from exc
    finally:
        engine.dispose()


def _build_file_preview(*, datasource, table_name: str | None, page: int, page_size: int) -> dict[str, Any]:
    storage_path = getattr(datasource, "storage_path", None)
    if not storage_path:
        raise ValueError("Datasource has no storage_path")

    file_type = normalize_datasource_type(getattr(datasource, "type", None))
    validate_file_type(file_type)

    raw = download_bytes(settings.MINIO_DATA_BUCKET, storage_path)
    if file_type == "csv":
        preview = _preview_csv(datasource=datasource, raw=raw, page=page, page_size=page_size)
    elif file_type == "json":
        preview = _preview_json(datasource=datasource, raw=raw, page=page, page_size=page_size)
    elif file_type == "parquet":
        preview = _preview_parquet(datasource=datasource, raw=raw, page=page, page_size=page_size)
    elif file_type == "xlsx":
        preview = _preview_xlsx(datasource=datasource, raw=raw, table_name=table_name, page=page, page_size=page_size)
    elif file_type == "xls":
        preview = _preview_xls(datasource=datasource, raw=raw, table_name=table_name, page=page, page_size=page_size)
    else:  # pragma: no cover - guarded by validate_file_type
        raise ValueError(f"Unsupported file type: {file_type}")
    return preview


def _preview_csv(*, datasource, raw: bytes, page: int, page_size: int) -> dict[str, Any]:
    table_name = _default_file_table_name(datasource)
    metadata_columns = _metadata_columns(datasource)
    offset = (page - 1) * page_size
    rows: list[dict[str, Any]] = []
    total_rows = 0

    stream = io.TextIOWrapper(io.BytesIO(raw), encoding="utf-8-sig", errors="replace", newline="")
    reader = csv.DictReader(stream)
    columns = reader.fieldnames or [column["name"] for column in metadata_columns]
    for index, row in enumerate(reader):
        if offset <= index < offset + page_size:
            rows.append(
                {
                    key: _serialize_preview_value(None if value == "" else value)
                    for key, value in row.items()
                }
            )
        total_rows += 1

    return _preview_payload(
        datasource=datasource,
        tables=[{"name": table_name}],
        selected_table=table_name,
        columns=_merge_columns(columns, metadata_columns, rows),
        rows=rows,
        page=page,
        page_size=page_size,
        total_rows=total_rows,
    )


def _preview_json(*, datasource, raw: bytes, page: int, page_size: int) -> dict[str, Any]:
    table_name = _default_file_table_name(datasource)
    text = raw.decode("utf-8-sig", errors="replace").strip()
    offset = (page - 1) * page_size
    rows: list[dict[str, Any]] = []
    total_rows = 0
    metadata_columns = _metadata_columns(datasource)

    if not text:
        return _preview_payload(
            datasource=datasource,
            tables=[{"name": table_name}],
            selected_table=table_name,
            columns=metadata_columns,
            rows=[],
            page=page,
            page_size=page_size,
            total_rows=0,
        )

    if text.startswith("[") or text.startswith("{"):
        parsed = json.loads(text)
        records = _normalize_json_records(parsed)
        total_rows = len(records)
        rows = [_serialize_preview_row(record) for record in records[offset : offset + page_size]]
        return _preview_payload(
            datasource=datasource,
            tables=[{"name": table_name}],
            selected_table=table_name,
            columns=_merge_columns(None, metadata_columns, rows),
            rows=rows,
            page=page,
            page_size=page_size,
            total_rows=total_rows,
        )

    for index, line in enumerate(text.splitlines()):
        stripped = line.strip()
        if not stripped:
            continue
        record = _normalize_json_value(json.loads(stripped))
        if offset <= total_rows < offset + page_size:
            rows.append(_serialize_preview_row(record))
        total_rows += 1

    return _preview_payload(
        datasource=datasource,
        tables=[{"name": table_name}],
        selected_table=table_name,
        columns=_merge_columns(None, metadata_columns, rows),
        rows=rows,
        page=page,
        page_size=page_size,
        total_rows=total_rows,
    )


def _preview_parquet(*, datasource, raw: bytes, page: int, page_size: int) -> dict[str, Any]:
    table_name = _default_file_table_name(datasource)
    parquet_file = pq.ParquetFile(io.BytesIO(raw))
    total_rows = int(parquet_file.metadata.num_rows if parquet_file.metadata else 0)
    offset = (page - 1) * page_size
    remaining = page_size
    skipped = 0
    rows: list[dict[str, Any]] = []

    for batch in parquet_file.iter_batches(batch_size=max(page_size, 256)):
        batch_rows = batch.num_rows
        if skipped + batch_rows <= offset:
            skipped += batch_rows
            continue
        batch_start = max(0, offset - skipped)
        take = min(batch_rows - batch_start, remaining)
        if take > 0:
            rows.extend(_serialize_preview_row(row) for row in batch.slice(batch_start, take).to_pylist())
            remaining -= take
        skipped += batch_rows
        if remaining <= 0:
            break

    columns = [{"name": field.name, "type": str(field.type)} for field in parquet_file.schema_arrow]
    return _preview_payload(
        datasource=datasource,
        tables=[{"name": table_name}],
        selected_table=table_name,
        columns=columns,
        rows=rows,
        page=page,
        page_size=page_size,
        total_rows=total_rows,
    )


def _preview_xlsx(*, datasource, raw: bytes, table_name: str | None, page: int, page_size: int) -> dict[str, Any]:
    workbook = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames
        selected_table = table_name.strip() if table_name and table_name.strip() in sheet_names else (sheet_names[0] if sheet_names else None)
        if not selected_table:
            return _preview_payload(
                datasource=datasource,
                tables=[],
                selected_table=None,
                columns=[],
                rows=[],
                page=page,
                page_size=page_size,
                total_rows=0,
            )

        sheet = workbook[selected_table]
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        columns_order = _normalize_headers(header)
        offset = (page - 1) * page_size
        rows: list[dict[str, Any]] = []
        total_rows = 0

        for index, values in enumerate(iterator):
            if offset <= index < offset + page_size:
                rows.append(_values_to_row(columns_order, values))
            total_rows += 1

        return _preview_payload(
            datasource=datasource,
            tables=[{"name": name} for name in sheet_names],
            selected_table=selected_table,
            columns=_merge_columns(columns_order, _metadata_columns(datasource), rows),
            rows=rows,
            page=page,
            page_size=page_size,
            total_rows=total_rows,
        )
    finally:
        workbook.close()


def _preview_xls(*, datasource, raw: bytes, table_name: str | None, page: int, page_size: int) -> dict[str, Any]:
    excel_file = pd.ExcelFile(io.BytesIO(raw))
    try:
        sheet_names = excel_file.sheet_names
        selected_table = table_name.strip() if table_name and table_name.strip() in sheet_names else (sheet_names[0] if sheet_names else None)
        if not selected_table:
            return _preview_payload(
                datasource=datasource,
                tables=[],
                selected_table=None,
                columns=[],
                rows=[],
                page=page,
                page_size=page_size,
                total_rows=0,
            )

        dataframe = excel_file.parse(sheet_name=selected_table)
        total_rows = len(dataframe.index)
        offset = (page - 1) * page_size
        page_frame = dataframe.iloc[offset : offset + page_size]
        rows = [_serialize_preview_row(record) for record in page_frame.to_dict(orient="records")]
        columns = [{"name": str(column), "type": str(dtype)} for column, dtype in dataframe.dtypes.items()]

        return _preview_payload(
            datasource=datasource,
            tables=[{"name": name} for name in sheet_names],
            selected_table=selected_table,
            columns=columns,
            rows=rows,
            page=page,
            page_size=page_size,
            total_rows=total_rows,
        )
    finally:
        excel_file.close()


def _preview_payload(
    *,
    datasource,
    tables: list[dict[str, Any]],
    selected_table: str | None,
    columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
    page: int,
    page_size: int,
    total_rows: int,
) -> dict[str, Any]:
    total_pages = 0 if not tables else max(1, math.ceil(total_rows / page_size))
    return {
        "datasource_id": getattr(datasource, "id"),
        "datasource_name": getattr(datasource, "name"),
        "category": getattr(datasource, "category", "database"),
        "tables": tables,
        "table": selected_table,
        "columns": columns,
        "rows": rows,
        "page": page,
        "page_size": page_size,
        "total_rows": total_rows,
        "total_pages": total_pages,
    }


def _default_file_table_name(datasource) -> str:
    return str(getattr(datasource, "name", None) or "data")


def _metadata_columns(datasource) -> list[dict[str, str]]:
    raw_columns = getattr(datasource, "file_metadata", None) or {}
    columns = raw_columns.get("columns", []) if isinstance(raw_columns, dict) else []
    result: list[dict[str, str]] = []
    for column in columns:
        if not isinstance(column, dict):
            continue
        name = str(column.get("name", "")).strip()
        if not name:
            continue
        result.append({"name": name, "type": str(column.get("type", "") or "")})
    return result


def _merge_columns(
    column_names: list[str] | None,
    metadata_columns: list[dict[str, str]],
    rows: list[dict[str, Any]],
) -> list[dict[str, str]]:
    ordered_names: list[str] = []
    seen: set[str] = set()
    for column in metadata_columns:
        name = column["name"]
        if name not in seen:
            seen.add(name)
            ordered_names.append(name)
    for name in column_names or []:
        if name not in seen:
            seen.add(name)
            ordered_names.append(name)
    for row in rows:
        for name in row.keys():
            if name not in seen:
                seen.add(name)
                ordered_names.append(name)

    metadata_by_name = {column["name"]: column.get("type", "") for column in metadata_columns}
    return [
        {
            "name": name,
            "type": metadata_by_name.get(name) or _infer_column_type(name, rows),
        }
        for name in ordered_names
    ]


def _infer_column_type(name: str, rows: list[dict[str, Any]]) -> str:
    for row in rows:
        value = row.get(name)
        if value is None:
            continue
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, int):
            return "integer"
        if isinstance(value, float):
            return "number"
        if isinstance(value, list):
            return "array"
        if isinstance(value, dict):
            return "object"
        return type(value).__name__
    return ""


def _normalize_headers(header: Any) -> list[str]:
    if header is None:
        return []
    raw_headers = list(header) if isinstance(header, (list, tuple)) else [header]
    normalized: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        base = str(value).strip() or f"column_{index + 1}"
        counts[base] = counts.get(base, 0) + 1
        suffix = counts[base]
        normalized.append(base if suffix == 1 else f"{base}_{suffix}")
    return normalized


def _values_to_row(columns: list[str], values: Any) -> dict[str, Any]:
    values_list = list(values) if isinstance(values, (list, tuple)) else [values]
    if len(values_list) < len(columns):
        values_list.extend([None] * (len(columns) - len(values_list)))
    return {
        column: _serialize_preview_value(values_list[index] if index < len(values_list) else None)
        for index, column in enumerate(columns)
    }


def _normalize_json_records(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [_normalize_json_value(item) for item in value]
    return [_normalize_json_value(value)]


def _normalize_json_value(value: Any) -> dict[str, Any]:
    if isinstance(value, dict):
        return value
    return {"value": value}


def _serialize_preview_row(row: dict[str, Any]) -> dict[str, Any]:
    return {str(key): _serialize_preview_value(value) for key, value in row.items()}


def _serialize_preview_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and math.isnan(value):
        return None
    if hasattr(value, "item"):
        try:
            return _serialize_preview_value(value.item())
        except Exception:
            return str(value)
    if isinstance(value, list):
        return [_serialize_preview_value(item) for item in value]
    if isinstance(value, tuple):
        return [_serialize_preview_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): _serialize_preview_value(item) for key, item in value.items()}
    return value
